from openpyxl_templates import TemplatedWorkbook
from openpyxl_templates.table_sheet import TableSheet
from openpyxl_templates.table_sheet.columns import CharColumn, IntColumn
from openpyxl_templates.templated_workbook import SheetnamesNotUnique
from openpyxl_templates.styles import DefaultStyleSet, StyleSet, ExtendedStyle, NamedStyle
from openpyxl.styles import Font, cell_style
#from openpyxl_templates.columns import RowStyle
from openpyxl_templates.utils import Typed, FakeCell
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting import Rule
#from style_xsds import xsd_style, comic_ext_style, cambria_style, xsd_c_style
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font





xsd_c_style = NamedStyle(
    name="xsd_c_style",
    font = Font(name='Calibri',
                    size=8,
                    bold=False,
                    italic=False,
                    vertAlign='baseline',#None,{‘superscript’, ‘subscript’, ‘baseline’}
                    underline='none',#'u'
                    strike=False,
                    color='FF000000'),
    fill = PatternFill(fill_type='solid',#None,#{'darkTrellis', 'lightUp', 'darkDown', 'darkGrid', 'lightGrid', 'lightTrellis', 'lightVertical', 'solid', 'gray125', 'lightGray', 'lightHorizontal', 'mediumGray', 'darkHorizontal', 'darkGray', 'lightDown', 'gray0625', 'darkVertical', 'darkUp'}
                    start_color='FFD9D9D9',
                    end_color='FF000000'),
    border = Border(left=Side(border_style='thin',#None,
                            color='FF000000'),
                    right=Side(border_style='thin',#None,
                            color='FF000000'),
                    top=Side(border_style='thick',#None,
                            color='FF000000'),
                    bottom=Side(border_style='thick',#None,
                                color='FF000000'),
                    diagonal=Side(border_style=None,
                                color='FF000000'),
                    diagonal_direction=0,
                    outline=Side(border_style=None,
                                color='FF000000'),
                    vertical=Side(border_style=None,
                                color='FF000000'),
                    horizontal=Side(border_style=None,
                                color='FF000000')
                ),
    alignment=Alignment(horizontal='center',#'general',
                        vertical='center',#'bottom',
                        text_rotation=0,
                        wrap_text=True,#False,
                        shrink_to_fit=True,#False,
                        indent=0),
    number_format = 'General',
    protection = Protection(locked=True,
                            hidden=False)
)





xsd_h_style = ExtendedStyle(
    base="xsd_c_style",
    name="xsd_h_style",
    font={
        "color": "FFFF0000",
    },
    fill=PatternFill(fill_type='solid',#None,#{'darkTrellis', 'lightUp', 'darkDown', 'darkGrid', 'lightGrid', 'lightTrellis', 'lightVertical', 'solid', 'gray125', 'lightGray', 'lightHorizontal', 'mediumGray', 'darkHorizontal', 'darkGray', 'lightDown', 'gray0625', 'darkVertical', 'darkUp'}
                    start_color='FF0000FF',
                    end_color='FF000000')
)

xsd_r_style = ExtendedStyle(
    base="xsd_c_style",
    name="xsd_r_style",
    font={
        "color": "FF0000FF"
    }
)


xsd_style = DefaultStyleSet(
    xsd_c_style,
    xsd_h_style,
    xsd_r_style
)






h_style=xsd_h_style#comic_ext_style
c_style=xsd_r_style#cambria_style
     
     
     
class ExcelStyleTemplate:
    def __init__(self) -> None:
        pass
    
    def add_column(self, )

     
     
     
     
        
class PrettyDemoSheet(TableSheet):
    def __init__(self,sheetname=None):
        #print("PRE SUPER INIT")
        super().__init__(
            sheetname,
            title_style="Default",#demo_style.__getitem__("Defaulta"),#"Bold & red, title",
            description_style="Extra tiny, description"#,
            #row_styles="red, title"
            #row_styles=hop_style
        )
        #comic_ext_style=comic_ext_style
        #cambria_style=cambria_style
        #print("POS SUPER INIT")
    #title_style="Bold & red, title",
    #description_style="Extra tiny, description"

    Element_Name = CharColumn(header_style=h_style,cell_style=c_style)#header_style="Comic_Style")#comic_style,cell_style=comic_ext_style)
    Data_Type = CharColumn(header_style=h_style,cell_style=c_style)#row_styles=comic_style)#header_style=h_style,cell_style=comic_style)
    Min_Occurs=CharColumn(header_style=h_style,cell_style=c_style)
    Max_Occurs=CharColumn(header_style=h_style,cell_style=c_style)
    Element_Description=CharColumn(header_style=h_style,cell_style=c_style)
    Table=CharColumn(header_style=h_style,cell_style=c_style)
    App_Column_Name=CharColumn(header_style=h_style,cell_style=c_style)
    #App_Name=CharColumn(header_style=h_style,cell_style=c_style)
    App_Data_Type=CharColumn(header_style=h_style,cell_style=c_style)
    App_Data_Length=CharColumn(header_style=h_style,cell_style=c_style)
    App_Column_Required=CharColumn(header_style=h_style,cell_style=c_style)
    App_Description=CharColumn(header_style=h_style,cell_style=c_style)

class DemoTemplatedWorkbook(TemplatedWorkbook):
    Sopapinho = PrettyDemoSheet(sheetname="AddrDesc")
    Sapopoinh = PrettyDemoSheet(sheetname="shashsd")
    pass
    
wb = DemoTemplatedWorkbook(data_only=False,
    file='C:\\Users\\ljdasilv\\Documents\\DBMS\\HTML_table_extract\\hopa.xlsx',template_styles=xsd_style)#template_styles=opa_style)
xx=[]
for ad in wb.Sopapinho:
    x=(ad)
    xx.append(x)
wb.Sapopoinh.write(
    objects=xx
)
wb.save("read_write34.xlsx")
 



