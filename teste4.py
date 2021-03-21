#%%
from openpyxl_templates import TemplatedWorkbook
from openpyxl_templates.table_sheet import TableSheet
from openpyxl_templates.table_sheet.columns import CharColumn, DEFAULT_COLUMN_WIDTH, IntColumn
from openpyxl_templates.templated_workbook import SheetnamesNotUnique
from openpyxl_templates.styles import DefaultStyleSet, StyleSet, ExtendedStyle, NamedStyle
from openpyxl.styles import Font, cell_style
#from openpyxl_templates.columns import RowStyle
from openpyxl_templates.utils import Typed, FakeCell
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting import Rule
from typing import Iterable, Tuple, TypeVar

#from style_xsds import xsd_style, comic_ext_style, cambria_style, xsd_c_style
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

from openpyxl import load_workbook, Workbook
import os
import sys
from copy import copy
from openpyxl.utils.cell import get_column_letter, column_index_from_string
from openpyxl.worksheet.dimensions import ColumnDimension, RowDimension, SheetDimension, SheetFormatProperties

DEFAULT_ROW_HEIGHT = 15
#MINIMUM_ROW_HEIGHT = 5
#MINIMUM_COLUMN_WIDTH = 5
#noNones = lambda fn, *args : fn(a for a in args if a is not None)
def noNones(fn, *args, default):
    try:
        return fn(a for a in args if a is not None)
    except:
        return default


def createNewWorkbook(manyWb):
    i=0
    for wb in manyWb:
        for sheetName in wb.sheetnames:
            o = theOne.create_sheet(sheetName)
            safeTitle = o.title
            copySheet(wb[sheetName],theOne[safeTitle])

def copyWorksheets(manyWb):
    i=0
    for wb in manyWb:
        for sheetName in wb.sheetnames:
            copySheet(wb[sheetName],theOne['Sheet'])
           # copySheet(wb[sheetName],theOne[sheetName+"_copy"])


def copySheet(sourceSheet,newSheet):
    i=0
    dims = {}
    rows = {}
    cols = {}
    #print(dir(sourceSheet.rows))
    #print(type(sourceSheet.rows))
    #j=len(sourceSheet.rows)
    r=0
    c=0
    global empty_sheet
    if empty_sheet:
        empty_sheet=False
    else:
        merged_cells_ranges = newSheet.merged_cells.ranges
        for merged_cell_range in merged_cells_ranges:
            #print("merged_cell:")
            #print(merged_cell_range)
            merged_cell_range.shift(row_shift=sourceSheet.max_row+1,col_shift=0)
        newSheet.insert_rows(idx=1,amount=sourceSheet.max_row+1)
        #k=0
        """
        for row in sourceSheet.rows:
            #if k==0:
                #print("row props:")
                #print(dir(row))
                #k+=1
            r+=1
            #rows[row] = rows.get(cell.row, 0)
        k=0
        for col in sourceSheet.columns:
            #cols[get_column_letter(col)] = cols.get(cell.column_letter, 0)
            if k==0:
                print("column:")
                print(col)
                print("dir col:")
                print(dir(col))
                k+=1
            c+=1
            #col.width=DEFAULT_COLUMN_WIDTH+1
            #print("col.width:")
            #print(col.width)
        """
        #newSheet.move_range('A1:Z1000',rows=j+2,cols=0,translate=True)
    print("sourceSheet:")
    print(sourceSheet)
    #print(dir(sourceSheet))
    #print(len(sourceSheet[get_column_letter(sourceSheet.max_column)]))
    #print(sourceSheet[get_column_letter(sourceSheet.max_column)])
    #print(len(sourceSheet[sourceSheet.max_row]))
    #print(sourceSheet[sourceSheet.max_row])
    
    #print("rows:")
    #print(j)
    """
    print("sheet.max_row:")
    print(sourceSheet.max_row)
    print("sheet.max_column:")
    print(sourceSheet.max_column)
    print("func number_of_rows:")
    r,c=get_max_row_with_col(sourceSheet)
    print("rows: "+str(r)+"  cols: "+str(c))
    print("func number_of_cols:")
    r,c=get_max_col_with_row(sourceSheet)
    print("rows: "+str(r)+"  cols: "+str(c))
    """
    for row in sourceSheet.rows:
        i+=1
        for cell in row:
            #print(dir(cell))
            if cell.value:
                #dims[cell.column] = max((dims.get(cell.column, 0), len(str(cell.dim_item))))
                 #dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.dim_item))))
                 """
                 print("cell:")
                 print(cell)
                 print("cell.row:")
                 print(cell.row)
                 print("cell.column_letter")
                 print(cell.column_letter)
                 """
                 #rows[cell.row] = rows.get(cell.row, 0)
                 #print(rows[cell.row])
                 #cols[cell.column_letter] = cols.get(cell.column_letter, 0)
                 #print(cols[cell.column_letter])
            newCell = newSheet.cell(row=cell.row, column=cell.column,#cell.col_idx,
                    value= cell.value)
            if cell.has_style:
                newCell.font = copy(cell.font)
                newCell.border = copy(cell.border)
                newCell.fill = copy(cell.fill)
                newCell.number_format = copy(cell.number_format)
                newCell.protection = copy(cell.protection)
                #newCell.alignment = cell.alignment.copy(shrink_to_fit=True,mergeCell=True,wrap_text=True,vertical='top')
                newCell.alignment = cell.alignment + Alignment(shrink_to_fit=True,mergeCell=True,wrap_text=True,vertical='top')
                #Alignment(shrink_to_fit=True,mergeCell=True,wrap_text=True,vertical='top')
                #newCell.offset = copy(cell.offset)
                #newCell.coordinate = copy(cell.coordinate)
                #newCell.
        #for col, dim_item in cols.items():
    #print("column_dimensions:")
    #print(sourceSheet.column_dimensions)
    #print("column_dimensions:")
    for col, dim_item in sourceSheet.column_dimensions.items():
        #print("col: ")
        #print(col)
        #print("dim_item:")
        #print(dim_item)
        #print(dir(dim_item))
        #newSheet.column_dimensions[col].width = min(50,dim_item)#min(50,dim_item)
        #newSheet.column_dimensions[col].width = min(50,sourceSheet.column_dimensions[col].width)
        #ccc = ColumnDimension(sourceSheet,col)
        min_max_delta = max(0,min(dim_item.max,sourceSheet.max_column) - dim_item.min)
        """
        print(sourceSheet)
        print("dim_item.max:")
        print(dim_item.max)
        print("dim_item.min:")
        print(dim_item.min)
        print("sourceSheet.max_column:")
        print(sourceSheet.max_column)
        print("min_max_delta:")
        print(min_max_delta)
        print("col:")
        print(col)
        print("colindex:")
        print(column_index_from_string(col))
        print("get_column_letter(column_index_from_string(col)+min_max_delta):")
        print(get_column_letter(column_index_from_string(col)+min_max_delta))
        """
        while True:
            next_col_idx = column_index_from_string(col)+min_max_delta
            next_col= get_column_letter(next_col_idx)
            if next_col_idx >= newSheet.max_column:
                newSheet.column_dimensions[next_col].min = dim_item.min+min_max_delta
                newSheet.column_dimensions[next_col].max = dim_item.min+min_max_delta
            newSheet.column_dimensions[next_col].width = noNones(max,dim_item.width,newSheet.column_dimensions[next_col].width,default=DEFAULT_COLUMN_WIDTH)
            #newSheet.column_dimensions[next_col].customWidth = dim_item.customWidth
            newSheet.column_dimensions[next_col].bestFit = dim_item.bestFit or newSheet.column_dimensions[next_col].bestFit
            #newSheet.column_dimensions[next_col].auto_size = dim_item.auto_size #alias for bestFit
            newSheet.column_dimensions[next_col].collapsed = dim_item.collapsed or newSheet.column_dimensions[next_col].collapsed
            newSheet.column_dimensions[next_col].hidden = dim_item.hidden or newSheet.column_dimensions[next_col].hidden
            newSheet.column_dimensions[next_col].outlineLevel = noNones(min,dim_item.outlineLevel,newSheet.column_dimensions[next_col].outlineLevel,default=0)
            #newSheet.column_dimensions[next_col].style = dim_item.style
            if min_max_delta == 0:
                break
            min_max_delta -= 1
    #for row, dim_item in rows.items():
    
    #print("\n\nmerged_cell_ranges:")
    for range in sourceSheet.merged_cells.ranges:
        #print(i)
        #print(dir(i))
        #flag_index = False
        #s = str(i).strip(string.uppercase)
        #if ... flag = True
        #if flag_index:
        #r1, r2, c1, c2 = range.min_row, range.max_row, range.min_col, range.max_col
        newSheet.merge_cells(range_string=str(range))#(start_row = r1, end_row = r2, start_column = c1, end_column = c2)

    for row, dim_item in sourceSheet.row_dimensions.items():
        #print(row)
        #newSheet.row_dimensions[row].height = min(50,dim_item)
        #newSheet.row_dimensions[row].height = min(50,sourceSheet.row_dimensions[row].height)
        #rrr = RowDimension(sourceSheet,row)
        #print("rrr:")
        #print(dir(rrr))
        #print(rrr.height)
        #print(rrr.ht)
        #if rrr.ht is None:
        #    rrr.ht = 14.50
        newSheet.row_dimensions[row].height = noNones(max,dim_item.height,newSheet.row_dimensions[row].height,default=DEFAULT_ROW_HEIGHT)
        newSheet.row_dimensions[row].collapsed = dim_item.collapsed or newSheet.row_dimensions[row].collapsed
        newSheet.row_dimensions[row].hidden = dim_item.hidden or newSheet.row_dimensions[row].hidden
        newSheet.row_dimensions[row].outlineLevel = noNones(min,dim_item.outlineLevel,newSheet.row_dimensions[row].outlineLevel,default=0)
        newSheet.row_dimensions[row].thickTop = dim_item.thickTop or newSheet.row_dimensions[row].thickTop
        newSheet.row_dimensions[row].thickBot = dim_item.thickBot or newSheet.row_dimensions[row].thickBot
        
    #newSheet.move_range('A1:Z1000',rows=i,cols=0,translate=True)

def get_col_length(sheet_obj,col):
    print(dir(sheet_obj))
    #return len(sheet_obj
  
def get_max_row_with_col(sheet_obj):
    number_of_rows = sheet_obj.max_row
    last_row_col = sheet_obj.max_column
    while number_of_rows > 0:
        col = last_row_col
        while col > 0:
            _cell = sheet_obj.cell(number_of_rows, col)
            if _cell.value != None or _cell.has_style:
                break
            else:
                col -= 1
        last_row_col = col
        if col == 0:
            number_of_rows -= 1
        else:
            break
    return (number_of_rows, last_row_col)

def get_max_col_with_row(sheet_obj):
    number_of_cols = sheet_obj.max_column
    last_col_row = sheet_obj.max_row
    while number_of_cols > 0:
        row = last_col_row
        while row > 0:
            _cell = sheet_obj.cell(row, number_of_cols)
            if _cell.value != None or _cell.has_style:
                break
            else:
                row -= 1
        last_col_row = row
        if row == 0:
            number_of_cols -= 1
        else:
            break
    return (last_col_row, number_of_cols)



#filesInput = ['hopa.xlsx','modelos\\mapeamento_header.xlsx','reda_abuera7.xlsx']#sys.argv[1:]
filesInput = ['modelos\\mapeamento_header.xlsx','modelos\\capa.xlsx','reda_abuera20.xlsx']#sys.argv[1:]
#filesInput = ['modelos\\mapeamento_header.xlsx','reda_abuera18.xlsx']#sys.argv[1:]
theOneFile = filesInput.pop(-1)
myfriends = [ load_workbook(f) for f in filesInput ]

#try this if you are bored
#myfriends = [ openpyxl.load_workbook(f) for k in range(200) for f in filesInput ]

theOne = Workbook()
empty_sheet=True
#print(dir(theOne['Sheet']))

#del theOne['Sheet'] #We want our new book to be empty. Thanks.
#createNewWorkbook(myfriends)
copyWorksheets(manyWb=myfriends)
#theOne['Sheet'].move_range(cell_range='A1:G10',rows=8,translate=True)
#theOne['Sheet'].insert_rows(idx=1,amount=8)
theOne.save(theOneFile)
