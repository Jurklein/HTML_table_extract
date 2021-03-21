#%%
from openpyxl_templates import TemplatedWorkbook
from openpyxl_templates.table_sheet import TableSheet
from openpyxl_templates.table_sheet.columns import CharColumn, IntColumn
from openpyxl_templates.templated_workbook import SheetnamesNotUnique
from openpyxl_templates.styles import DefaultStyleSet, StyleSet, ExtendedStyle, NamedStyle
from openpyxl.styles import Font, cell_style
#from openpyxl_templates.columns import RowStyle
from openpyxl_templates.utils import Typed, FakeCell
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting import Rule
from typing import Iterable, Tuple, TypeVar

#from style_xsds import xsd_style, comic_ext_style, cambria_style, xsd_c_style
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

from openpyxl import load_workbook, Workbook
import os
import sys
from copy import copy

"""
fwb = Workbook()

wb = load_workbook('modelos\\capa.xlsx')
ws1 = wb.active
wb = load_workbook('modelos\\mapeamento_header.xlsx')
ws2 = wb.active
wb = load_workbook('modelos\\overview_funcional.xlsx')
ws3 = wb.active
wb = load_workbook('modelos\\regras_de_negocio.xlsx')
ws4 = wb.active


#fwb.save('reda_wraita4.xlsx')
"""

def createNewWorkbook(manyWb):
    for wb in manyWb:
        for sheetName in wb.sheetnames:
            o = theOne.create_sheet(sheetName)
            safeTitle = o.title
            copySheet(wb[sheetName],theOne[safeTitle])

def copySheet(sourceSheet,newSheet):
    for row in sourceSheet.rows:
        for cell in row:
            newCell = newSheet.cell(row=cell.row, column=cell.col_idx,
                    value= cell.value)
            if cell.has_style:
                newCell.font = copy(cell.font)
                newCell.border = copy(cell.border)
                newCell.fill = copy(cell.fill)
                newCell.number_format = copy(cell.number_format)
                newCell.protection = copy(cell.protection)
                newCell.alignment = copy(cell.alignment)

filesInput = ['modelos\\capa.xlsx','modelos\\mapeamento_header.xlsx','reda_abuera.xlsx']#sys.argv[1:]
theOneFile = filesInput.pop(-1)
myfriends = [ load_workbook(f) for f in filesInput ]

#try this if you are bored
#myfriends = [ openpyxl.load_workbook(f) for k in range(200) for f in filesInput ]

theOne = Workbook()
del theOne['Sheet'] #We want our new book to be empty. Thanks.
createNewWorkbook(myfriends)
theOne.save(theOneFile)


#%%
class DemoTableSheet(TableSheet):
    column1 = CharColumn()
    columnaa2 = IntColumn()


# ------------------- Styling -------------------
#_T = TypeVar("_T")

class RowNamedStyle(NamedStyle):
    cell_style = Typed("cell_style", expected_types=[str, NamedStyle, ExtendedStyle], allow_none=True)
    data_validation = Typed("data_validation", expected_type=DataValidation, allow_none=True)
    conditional_formatting = Typed("conditional_formatting", expected_type=Rule, allow_none=True)

    def __init__(self, row_type='Header', getter=None, cell_style=None, conditional_formatting=None, data_validation=None,**kwargs):
        super(RowNamedStyle, self).__init__(**kwargs)
        self.row_type=row_type
        self.getter = getter
        self.cell_style = cell_style
        self.conditional_formatting = conditional_formatting
        self.data_validation = data_validation

    #def extend(self, __iterable: Iterable[_T]) -> None: ...
""" 
class RowNamedStyle(NamedStyle):
    def __init__(self, row_type='Header', getter=None,**kwargs):
        super(RowNamedStyle, self).__init__(**kwargs)
        self.__row_type = row_type
        self.getter=getter

    #row_type='Header'
    @property
    def row_type(self):
        return self.__row_type
    @row_type.setter
    def row_type(self, var='Header'):
            self.__row_type = var

 """
comic_style = NamedStyle(
    name="Comic_Style",
    font=Font(
        name="Comic Sans MS",
        size=6
    )#,
    #row_type='Header'
)


comic_ext_style = ExtendedStyle(
    base="Comic_Style",
    name="Comic_Ext_Style",
    font= {
        "bold": True,
    }
)

hop_style = DefaultStyleSet(
    comic_style,
    comic_ext_style
)

cambria_style = ExtendedStyle(
    base="Comic_Style",
    name="Cambria_Style",
    font=Font(
        name="Cambria",
        size=24
    )
)


opa_style = DefaultStyleSet(
    comic_style,
    cambria_style
)

demo_style = DefaultStyleSet(
    NamedStyle(
        name="Default",
        font=Font(
            name="Arial",
            size=12,
            bold=True
        )
    ),
    ExtendedStyle(
        base="Default", # Reference to the style defined above
        name="Header",
        font={
            "name": "Arial",
            "bold": True,
        }
    )
)



























xsd_c_style = NamedStyle(
    name="xsd_c_style",
    font = Font(name='Calibri',
                    size=8,
                    bold=False,
                    italic=False,
                    vertAlign='baseline',#None,{‘superscript’, ‘subscript’, ‘baseline’}
                    underline='none',#'u'
                    strike=False,
                    color='FF000000'),
    fill = PatternFill(fill_type='solid',#None,#{'darkTrellis', 'lightUp', 'darkDown', 'darkGrid', 'lightGrid', 'lightTrellis', 'lightVertical', 'solid', 'gray125', 'lightGray', 'lightHorizontal', 'mediumGray', 'darkHorizontal', 'darkGray', 'lightDown', 'gray0625', 'darkVertical', 'darkUp'}
                    start_color='FFD9D9D9',
                    end_color='FF000000'),
    border = Border(left=Side(border_style='thin',#None,
                            color='FF000000'),
                    right=Side(border_style='thin',#None,
                            color='FF000000'),
                    top=Side(border_style='thick',#None,
                            color='FF000000'),
                    bottom=Side(border_style='thick',#None,
                                color='FF000000'),
                    diagonal=Side(border_style=None,
                                color='FF000000'),
                    diagonal_direction=0,
                    outline=Side(border_style=None,
                                color='FF000000'),
                    vertical=Side(border_style=None,
                                color='FF000000'),
                    horizontal=Side(border_style=None,
                                color='FF000000')
                ),
    alignment=Alignment(horizontal='center',#'general',
                        vertical='center',#'bottom',
                        text_rotation=0,
                        wrap_text=True,#False,
                        shrink_to_fit=True,#False,
                        indent=0),
    number_format = 'General',
    protection = Protection(locked=True,
                            hidden=False)
)

"""
xsd_c_style.font = font
xsd_c_style.fill = fill
xsd_c_style.border = border
xsd_c_style.alignment = alignment
xsd_c_style.number_format = number_format
xsd_c_style.protection = protection
"""
"""
xsd_c_style = NamedStyle(
    name="xsd_c_style",
    font=Font(
        name="Comic Sans MS",
        size=6
    )#,
    #row_type='Header'
)
"""

xsd_h_style = ExtendedStyle(
    base="xsd_c_style",
    name="xsd_h_style",
    font={
        "color": "FFFF0000",
    },
    fill=PatternFill(fill_type='solid',#None,#{'darkTrellis', 'lightUp', 'darkDown', 'darkGrid', 'lightGrid', 'lightTrellis', 'lightVertical', 'solid', 'gray125', 'lightGray', 'lightHorizontal', 'mediumGray', 'darkHorizontal', 'darkGray', 'lightDown', 'gray0625', 'darkVertical', 'darkUp'}
                    start_color='FF0000FF',
                    end_color='FF000000')
)

xsd_r_style = ExtendedStyle(
    base="xsd_c_style",
    name="xsd_r_style",
    font={
        "color": "FF0000FF"
    }
    #,fill=PatternFill(fill_type='darkTrellis',#None,#{'darkTrellis', 'lightUp', 'darkDown', 'darkGrid', 'lightGrid', 'lightTrellis', 'lightVertical', 'solid', 'gray125', 'lightGray', 'lightHorizontal', 'mediumGray', 'darkHorizontal', 'darkGray', 'lightDown', 'gray0625', 'darkVertical', 'darkUp'}
    #                start_color='FF0000FF',
    #                end_color='FF000000')
)


xsd_style = DefaultStyleSet(
    xsd_c_style,
    xsd_h_style,
    xsd_r_style
)


#comic_ext_style=comic_ext_style
#cambria_style=cambria_style














h_style=xsd_h_style#comic_ext_style
c_style=xsd_r_style#cambria_style
        
class PrettyDemoSheet(TableSheet):
    lola={}
    
    def __init__(self,sheetname=None,h_style=None, c_style=None):
        #print("PRE SUPER INIT")
        super().__init__(
            sheetname,
            title_style="Default",#demo_style.__getitem__("Defaulta"),#"Bold & red, title",
            description_style="Extra tiny, description"#,
            #row_styles="red, title"
            #row_styles=hop_style
        )
        self.h_style=h_style
        self.c_style=c_style
        #self.lola=[]
        #self.columns
        #comic_ext_style=comic_ext_style
        #cambria_style=cambria_style
        #print("POS SUPER INIT")
    #title_style="Bold & red, title",
    #description_style="Extra tiny, description"
        #self.add_column(CharColumn("Element_Name"),object_attribute={"header_style":h_style,"cell_style":c_style})
        #self.columns.insert(0,CharColumn(header="Element_Name",header_style=h_style,cell_style=c_style,object_attribute="Element_Name"))
    Element_Name=CharColumn(header="Element_Name",header_style=h_style,cell_style=c_style,object_attribute="Element_Name")
    #i=CharColumn(header="Element_Name",header_style=h_style,cell_style=c_style)#header_style="Comic_Style")#comic_style,cell_style=comic_ext_style)
    Data_Type = CharColumn(header_style=h_style,cell_style=c_style)#row_styles=comic_style)#header_style=h_style,cell_style=comic_style)
    Min_Occurs=CharColumn(header_style=h_style,cell_style=c_style)
    Max_Occurs=CharColumn(header_style=h_style,cell_style=c_style)
    Element_Description=CharColumn(header_style=h_style,cell_style=c_style)
    Table=CharColumn(header_style=h_style,cell_style=c_style)
    App_Column_Name=CharColumn(header_style=h_style,cell_style=c_style)
    #App_Name=CharColumn(header_style=h_style,cell_style=c_style)
    App_Data_Type=CharColumn(header_style=h_style,cell_style=c_style)
    App_Data_Length=CharColumn(header_style=h_style,cell_style=c_style)
    App_Column_Required=CharColumn(header_style=h_style,cell_style=c_style)
    App_Description=CharColumn(header="App_Description",header_style=h_style,cell_style=c_style)

class DemoTemplatedWorkbook(TemplatedWorkbook):
    #demo_sheet1 = DemoTableSheet(sheetname="Opalorderzinho meum")
    #demo_sheet1 = PrettyDemoSheet(sheetname="Opalorderzinho meum")
    #demo_sheet2 = DemoTableSheet(sheetname="Opatapoptoptopinhodois")
    Data_Tbl = PrettyDemoSheet(sheetname="AddrDesc")#Sheetname MUST be equal to the base sheet name
    Capa = PrettyDemoSheet(sheetname="capa")
    Mapeamento_Header = PrettyDemoSheet(sheetname="mapeamento_header")
    
    pass
    
wb = DemoTemplatedWorkbook(data_only=True,
    file='C:\\Users\\ljdasilv\\Documents\\DBMS\\HTML_table_extract\\hopa.xlsx',template_styles=xsd_style)
#wb.Data_Tbl.add_column(column=CharColumn(header="App_Description",header_style=h_style,cell_style=c_style,object_attribute="App_Description"))
#print(wb.Sopapinho.columns)
#wb.Data_Tbl.columns[0].header
#print(wb.Sopapinho.App_Column_Name.object_attribute)
#print(wb.Sopapinho.App_Data_Length.object_attribute)
#wb.Sopapinho.columns[0].object_attribute="Element_Name"
#print(wb.Sopapinho.columns[0].object_attribute)# App_Data_Length.object_attribute)
#print("LOLAAAA:")
#wb.Sopapinho.lola[wb.Sopapinho.columns[0].header]=wb.Sopapinho.columns[0]
#print(dir(wb.Sopapinho.lola["Element_Name"]))
#print(wb.Sopapinho.lola["Element_Name"].header_style)
#print(wb.Sopapinho.App_Column_Name.header_style)
#print(wb.Sopapinho.Element_Name.header_style)
#Sapuquinho = PrettyDemoSheet(sheetname="Sapuquinho")
#wb.add_templated_sheet(Sapuquinho,sheetname="Sapuquinho")
#wb.add_templated_sheet(PrettyDemoSheet(),sheetname="Quarta")
#Opalinho = PrettyDemoSheet(sheetname="Opalinho")
#Sapuquinho.write(title="Opalele",objects=(("Ahea","asdasd","frf","","","","","","","",""),("asdasdsa","Asdasd","","","","","","","","","")))
#print("Sheetnames:")
#print(wb.sheetnames)
#print("Templated Sheets:")
#print(wb.templated_sheets)
#print("Sapuquinho::")
#print(wb.templated_sheets[3])
#print(wb.workbook["AddrDesc"].columns)

#print(wb.Sopapinho.lola.[0].header_style)
#print(wb.Sopapinho.lola[0])
#wb.workbook.add_named_style(comic_style)
"""print('wb.workbook.named_styles:')
print(wb.workbook.named_styles)
print('wb.template_styles.names:')
print(wb.template_styles.names)
print("wb.sheetnames:")
print(wb.sheetnames)
print("wb.templated_sheets:")
print(wb.templated_sheets)"""
#print(wb.AddrDesc.write(preserve=True,title="dxrdtr rtdrt"))
xx=[]
#print("VAI ENTRAR NO FORZERA")
#iteriter = wb.Sopapinho.__iter__()
#i=0
#for ad in wb.Sopapinho:

"""
wb.AddrDesc.write(
    preserve=True,
    objects=wb.AddrDesc,
    look_for_headers=True
)
"""

for ad in wb.Data_Tbl:
    #if i==0:
    #    print("entrou no forzera")
    #    print(ad)
    #    i+=1
    #print("\nITERITER:")
    #print(next(iteriter))
    #print("ad.locals:")
    #print(ad.locals())
    x=(ad)
    """
    #print(
            ad.Element_Name,
            ad.Data_Type,
            ad.Min_Occurs,
            ad.Max_Occurs,
            ad.Element_Description,
            ad.Table,
            ad.App_Column_Name,
            #ad.App_Name=CharColum
            ad.App_Data_Type,
            ad.App_Data_Length,
            ad.App_Column_Required,
            ad.App_Description)
            """
    xx.append(x)

#wb=DemoTemplatedWorkbook(template_styles=opa_style)
#wb.Sopapinho.write(
#wb.Sapopoinh.write(
#    objects=xx
    #title="AHoashdo"
#)

#wb.Sopapinho.remove()
#del wb.workbook['AddrDesc']
try:
    del wb.workbook['Sheet']
except:
    pass

#print(wb.workbook.sheetnames)

#wb.Sapopoinh.i.header_style=xsd_r_style
#wb.Sapopoinh.App_Column_Name.cell_style=xsd_h_style
#wb.Sapopoinh.write(

wb.Data_Tbl.write(
    preserve=False,
    objects=xx,
    look_for_headers=True
)
wb.save("read_write34.xlsx")

wb1=wb

wb=DemoTemplatedWorkbook(data_only=True,
    file='C:\\Users\\ljdasilv\\Documents\\DBMS\\HTML_table_extract\\modelos\\capa.xlsx',template_styles=xsd_style)
wb.Data_Tbl=wb1.Data_Tbl


wb.save("read_write4.xlsx")

"""
wb.Capa.write(
    preserve=False,
    objects=wb.Capa
)

wb.save("read_write5.xlsx")
"""


#cap=[]
#for ad in wb.Capa:
#    cap.append((ad))

#wb.Capa.write(
#    preserve=False,
#    objects=cap
#)

#wb.save("read_write6.xlsx")

    #print(ad.Element_Name, ad.Table, ad.App_Column_Name)

#opa=wb.add_templated_sheet(sheet=wb.AddrDesc,sheetname="ASDAHSD",add_to_self=True)
#print('wb.demo_sheet1.column1.header_style:')
#print(wb.demo_sheet1.column1.header_style)
#print('wb.demo_sheet1.row_styles:')
#print(wb.demo_sheet1.row_styles)
#wb.demo_sheet1.add_row_style(comic_style)
#print('wb.demo_sheet1.row_styles:')
#print(wb.demo_sheet1.row_styles)

#print(wb.demo_sheet1.column1.create_header(wb.demo_sheet1,opa_style))
#wb.demo_sheet1.column1.header_style#="Cambria_Style"#comic_ext_style#"Comic_Style"

"""
wb.demo_sheet1.write(
    objects=(
        ("Row 1", 1),
        ("Row 2", 2),
        ("Row 3", 3),
    ),
    title="The first sheet"
)

#wb = DemoTemplatedWorkbook(template_styles=opa_style)

wb.demo_sheet2.write(
    objects=(
        ("Row 1", 1),
        ("Row 2", 2),
        ("Row 3", 3),
    )#,
    #title="The second sheet",
    #description="Lorem ipsum dolor sit amet, consectetur adipiscing elit. In euismod, sem eu."
)
"""
#wb.workbook.remove_sheet('Sheet') #Deprecated
#wb.workbook.remove(wb.workbook.get_sheet_by_name('Sheet')) #get_sheet_by_name Deprecated
#wb.workbook.remove(wb.workbook['Sheet'])
#del wb.workbook['Sheet']


"""
INTERFACE ID
XSD Name
XSD Element Name
XSD Element Data Type
XSD Min Occurs
XSD Max Occurs
XSD Element Description
TABLE
COLUMN
TYPE
COL LENGTH
IS REQUIRED?
ORIGIN Description
SOURCE
TARGET
"""
wb.save("read_write.xlsx")
#wb = DemoTemplatedWorkbook("read_write.xlsx")
"""
print('dir(wb):')
print(dir(wb))
print('dir(wb.item_class):')
print(dir(wb.item_class))
print('wb.item_class:')
print(wb.item_class)
print('dir(wb.sheetnames):')
print(dir(wb.sheetnames))
print('wb.sheetnames:')
print(wb.sheetnames)
print('dir(wb.template_styles):')
print(dir(wb.template_styles))
print('wb.template_styles:')
print(wb.template_styles)
print('wb.template_styles.names:')
print(wb.template_styles.names)
print('dir(wb.templated_sheets):')
print(dir(wb.templated_sheets))
print('wb.templated_sheets:')
print(wb.templated_sheets)
print('dir(wb.workbook):')
print(dir(wb.workbook))
print('wb.workbook:')
print(wb.workbook)

"""
"""
wb = DemoTemplatedWorkbook("read_write.xlsx")

for row in wb.demo_sheet1.read():
    print(row)
for row in wb.demo_sheet2.read():
    print(row)
"""
"""
wb.demo_sheet1.write(
    objects=(
        ("Row 1", 1),
        ("Row 2", 2),
        ("Row 3", 3),
    ),
    title="The first sheet againnn"
)
wb.save("read_write_2.xlsx")

"""
# --------------- Passing Objects ---------------
"""
class DemoObject():
    def __init__(self, column1, column2):
        self.column1 = column1
        self.columnaa2 = column2


wb = DemoTemplatedWorkbook()
wb.demo_sheet1.write(
    objects=(
        DemoObject("Row 1", 1),
        DemoObject("Row 2", 2),
        DemoObject("Row 3", 3),
    )
)
"""

"""
# ------------------- Styling -------------------
class PrettyDemoSheet(TableSheet):
    def __init__(self):
        super().__init__(
            title_style="Bold & red, title",
            description_style="Extra tiny, description"#,
            #row_styles="red, title"
        )
    #title_style="Bold & red, title",
    #description_style="Extra tiny, description"
    column1 = CharColumn()
    columnaa2 = IntColumn()


# --------------- Reading ---------------
"""

"""
wb = DemoTemplatedWorkbook("read_write.xlsx")
wb.demo_sheet1 = PrettyDemoSheet(sheetname="Opalorduuerzinho meum")
wb.demo_sheet1.write(
    objects=(
        ("Row 1", 1),
        ("Row 2", 2),
        ("Row 3", 3),
    ),
    title="Opalelezinhomeu"
)
#for row in wb.demo_sheet1.read():
#    print(row)
wb.save("read_write_3.xlsx")

""" # --------------- As iterator ---------------
#for row in wb.demo_sheet2:
#    print(row) """